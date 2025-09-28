from telebot.async_telebot import AsyncTeleBot
from openpyxl import Workbook, load_workbook
from russian import test_rus
from telebot.types import *
from tajik import test_tjk
from datetime import *
import logging
import asyncio
import hashlib
import shutil
import json
import math
import os

logging.basicConfig(filename='bot.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
bot = AsyncTeleBot('8263239418:AAF--j6XK5lrsyLoyyJWB6bHq4dY9Ju1sEU')

user_lang = {}
user_mark = []
user_test_state = {}
password_waiting = {}
USERS_FILE = 'users.txt'
RESULTS_FILE = 'test_results.xlsx'
STATE_FILE = 'state.json'
BACKUP_DIR = 'backups'
ADMIN_ID = 7205513397
ADMIN3_ID = 1354151664
ADMIN4_ID = 5420002634
ADMINS = {ADMIN_ID, ADMIN3_ID, ADMIN4_ID}

PASSWORD_HASH ="softclub2050"
all_message_ids = {}

DELETION_QUEUE_FILE = 'deletion_queue.json'
deletion_queue = []  
last_activity = {}  
WATCHDOG_TIMEOUT_SECONDS = 600  
QUESTION_TTL_SECONDS = 24 * 3600  
TEST_TOTAL_LIMIT_SECONDS = 3600  


# ---------------------------------Фармонҳои бот-------------------------------------
async def set_commands():
    try:
        commands = [
            BotCommand(command="/start", description="Барои оғоз кардани кор бо бот 🚀"),
            BotCommand(command="/language", description="Барои иваз кардани забон 🌐"),
            BotCommand(command="/test", description="Барои оғоз кардани тест 📝"),
            BotCommand(command="/mark", description="Барои дидани баҳои гирифта 〽️"),
            BotCommand(command="/help", description="Барои гирифтани кумак ❓"),
            BotCommand(command="/restart_test", description="Барои супоридани такрори тест 📝"),
            BotCommand(command="/cancel_test", description="Бекоркунии фаврии тест ❌"),
        ]
        await bot.set_my_commands(commands)
        logging.info("Bot commands set successfully")
    except Exception as e:
        logging.error(f"Error setting commands: {e}")
# -----------------------------------------------------------------------------------

# ---- ADMIN HELPERS (ИЛОВА) ----
def is_admin(uid: int) -> bool:
    return uid in ADMINS
# --------------------------------


# ---------------------------------Фармони /kabir-------------------------------------
@bot.message_handler(commands=['admin'])
async def kabir(message):
    try:
        user_id = int(message.chat.id)
        if not is_admin(user_id):   
            msg = {
                'tj': "🚫 Шумо иҷозати истифодаи ин фармонро надоред!",
                'ru': "🚫 У вас нет прав для использования этой команды!"
            }.get(user_lang.get(user_id, 'ru'), 'ru')
            await bot.send_message(user_id, msg)
            return

        if not os.path.exists(RESULTS_FILE):
            await bot.send_message(user_id, "📂 Файли натиҷаҳо вуҷуд надорад!")
            return

        wb = load_workbook(RESULTS_FILE)
        ws = wb.active
        user_data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            user_data.append({
                'Имя': row[1] if row[1] else "N/A",
                'Фамилия': row[2] if row[2] else "N/A",
                'Username': row[3] if row[3] else "N/A",
                'Возраст': row[5] if row[5] else "N/A",
                'Баллы': row[6] if row[6] else 0,
            })

        if not user_data:
            await bot.send_message(user_id, "📭 Ягон маълумот дар файл мавҷуд нест!")
            return
        items_per_page = 5  
        total_pages = (len(user_data) + items_per_page - 1) // items_per_page
        current_page = 1

        async def send_page(page):
            start_idx = (page - 1) * items_per_page
            end_idx = start_idx + items_per_page
            page_data = user_data[start_idx:end_idx]

            message_text = "📊 **Маълумоти корбарон**:\n\n"
            message_text += "┌──────────────────────────────┐\n"
            for i, data in enumerate(page_data, start_idx + 1):
                message_text += (
                    f"│ #{i} 🧑‍💼\n"
                    f"│ Имя: {data['Имя']}\n"
                    f"│ Фамилия: {data['Фамилия']}\n"
                    f"│ Username: {data['Username']}\n"
                    f"│ Возраст: {data['Возраст']}\n"
                    f"│ Баллы: {data['Баллы']} 🎯\n"
                    f"└──────────────────────────────┘\n\n"
                )
            message_text += f"📄 Саҳифа: {page}/{total_pages} | Итого: {len(user_data)} корбар"
            markup = InlineKeyboardMarkup()
            if total_pages > 1:
                buttons = []
                if page > 1:
                    buttons.append(InlineKeyboardButton("⬅️ Саҳифаи пеш", callback_data=f"page_{page-1}"))
                if page < total_pages:
                    buttons.append(InlineKeyboardButton("Саҳифаи навбатӣ ➡️", callback_data=f"page_{page+1}"))
                markup.add(*buttons)
            await bot.send_message(user_id, message_text, parse_mode='Markdown', reply_markup=markup)
            logging.info(f"Sent page {page} of user data to admin {user_id}")
        await send_page(current_page)

    except Exception as e:
        logging.error(f"Error in kabir command: {e}")
        await bot.send_message(user_id, "❌ Хато рух дод. Лутфан дубора кӯшиш кунед.")


@bot.callback_query_handler(func=lambda call: call.data.startswith('page_'))
async def handle_page_change(call):
    try:
        user_id = call.from_user.id
        if user_id != ADMIN_ID:
            await bot.answer_callback_query(call.id, "Шумо иҷозат надоред!")
            return
        page = int(call.data.split('_')[1])
        wb = load_workbook(RESULTS_FILE)
        ws = wb.active
        user_data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            user_data.append({
                'Имя': row[1] if row[1] else "N/A",
                'Фамилия': row[2] if row[2] else "N/A",
                'Username': row[3] if row[3] else "N/A",
                'Возраст': row[5] if row[5] else "N/A",
                'Баллы': row[6] if row[6] else 0,
            })
        items_per_page = 5
        total_pages = (len(user_data) + items_per_page - 1) // items_per_page

        if 1 <= page <= total_pages:
            start_idx = (page - 1) * items_per_page
            end_idx = start_idx + items_per_page
            page_data = user_data[start_idx:end_idx]
            message_text = "📊 **Маълумоти корбарон**:\n\n"
            message_text += "┌──────────────────────────────┐\n"
            for i, data in enumerate(page_data, start_idx + 1):
                message_text += (
                    f"│ #{i} 🧑‍💼\n"
                    f"│ Имя: {data['Имя']}\n"
                    f"│ Фамилия: {data['Фамилия']}\n"
                    f"│ Username: {data['Username']}\n"
                    f"│ Возраст: {data['Возраст']}\n"
                    f"│ Баллы: {data['Баллы']} 🎯\n"
                    f"└──────────────────────────────┘\n\n"
                )
            message_text += f"📄 Саҳифа: {page}/{total_pages} | Итого: {len(user_data)} корбар"
            markup = InlineKeyboardMarkup()
            if total_pages > 1:
                buttons = []
                if page > 1:
                    buttons.append(InlineKeyboardButton("⬅️ Саҳифаи пеш", callback_data=f"page_{page-1}"))
                if page < total_pages:
                    buttons.append(InlineKeyboardButton("Саҳифаи навбатӣ ➡️", callback_data=f"page_{page+1}"))
                markup.add(*buttons)
            await bot.edit_message_text(
                chat_id=user_id,
                message_id=call.message.message_id,
                text=message_text,
                parse_mode='Markdown',
                reply_markup=markup
            )
            logging.info(f"Page {page} sent to admin {user_id}")
        await bot.answer_callback_query(call.id)

    except Exception as e:
        logging.error(f"Error in page change: {e}")
        await bot.answer_callback_query(call.id, "Хато рух дод!")
# ---------------------------------------------------------------------------------------



# ---------------------------------Идоракунии файлҳо-------------------------------------
def check_file_integrity(file_path):
    try:
        if not os.path.exists(file_path):
            with open(file_path, 'w', encoding='utf-8') as f:
                pass
        return True
    except Exception as e:
        logging.error(f"Error checking file {file_path}: {e}")
        return False

def create_backup(file_path):
    try:
        if not os.path.exists(BACKUP_DIR):
            os.makedirs(BACKUP_DIR)
        
        file_name = os.path.basename(file_path)
        backup_path = os.path.join(BACKUP_DIR, f"{file_name}.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
        
        shutil.copy(file_path, backup_path)
        logging.info(f"Backup created: {backup_path}")
    except Exception as e:
        logging.error(f"Error creating backup for {file_path}: {e}")
# ---------------------------------------------------------------------------------------



# ---------------------------------Захира ва барқарорсозӣ--------------------------------
def load_state():
    try:
        if os.path.exists(STATE_FILE):
            with open(STATE_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                global deletion_queue, last_activity
                deletion_queue = data.get('deletion_queue', [])

                last_activity = {int(k): datetime.fromisoformat(v) for k, v in data.get('last_activity', {}).items()}
                return (
                    data.get('user_lang', {}),
                    data.get('user_test_state', {}),
                    data.get('user_mark', []),
                    data.get('password_waiting', {})
                )
        return {}, {}, [], {}
    except Exception as e:
        logging.error(f"Error loading state: {e}")
        return {}, {}, [], {}

def save_state():
    try:
        with open(STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump({
                'user_lang': user_lang,
                'user_test_state': user_test_state,
                'user_mark': user_mark,
                'password_waiting': password_waiting,
                'deletion_queue': deletion_queue,
                'last_activity': {str(k): v.isoformat() for k, v in last_activity.items()},
            }, f, ensure_ascii=False, indent=2)
        logging.info("State saved successfully")
    except Exception as e:
        logging.error(f"Error saving state: {e}")

async def save_state_periodically():
    while True:
        save_state()
        await asyncio.sleep(60)

user_lang, user_test_state, user_mark, password_waiting = load_state()
# -----------------------------------------------------------------------------------


# ---------------------------------Нест кардани маълумот--------------------------------
def delete_user_from_excel(user_id, filename=RESULTS_FILE):
    try:
        if not os.path.exists(filename):
            logging.warning(f"File {filename} not found.")
            return
        create_backup(filename)
        wb = load_workbook(filename)
        ws = wb.active
        rows_to_delete = []
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(user_id):
                rows_to_delete.append(row[0].row)
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)
        wb.save(filename)
        logging.info(f"Deleted {len(rows_to_delete)} rows for user_id {user_id} from {filename}")
    except Exception as e:
        logging.error(f"Error deleting user from Excel: {e}")

@bot.message_handler(commands=['restart_test'])
async def ask_password(message):
    try:
        user_id = message.chat.id
        password_waiting[user_id] = True
        if user_id not in user_lang:
            user_lang[user_id] = 'ru'
        msg = {
            'tj': "🔐 Лутфан рамзи махсусро ворид кунед барои такроран ҳал кардани тестҳо:",
            'ru': "🔐 Пожалуйста, введите специальный пароль для повторного прохождения теста:"
        }.get(user_lang[user_id], 'ru')
        sent_msg = await bot.send_message(user_id, msg)
        if user_id not in all_message_ids:
            all_message_ids[user_id] = []
        all_message_ids[user_id].append(sent_msg.message_id)
        _track_for_deletion(user_id, sent_msg.message_id)
        logging.info(f"Asked password for restart_test from user {user_id}")
        save_state()
    except Exception as e:
        logging.error(f"Error in ask_password for user {user_id}: {e}")
        sent_msg = await bot.send_message(user_id, "Хато рух дод. Лутфан дубора кӯшиш кунед.")
        if user_id not in all_message_ids:
            all_message_ids[user_id] = []
        all_message_ids[user_id].append(sent_msg.message_id)
        _track_for_deletion(user_id, sent_msg.message_id)

@bot.message_handler(func=lambda msg: password_waiting.get(msg.chat.id, False))
async def check_password(message):
    try:
        user_id = message.chat.id
        user_input = message.text.strip()
        try:
            await bot.delete_message(user_id, message.message_id)
            logging.info(f"Deleted password message {message.message_id} for user {user_id}")
        except Exception as e:
            logging.error(f"Error deleting password message {message.message_id} for user {user_id}: {e}")
        if user_input == PASSWORD_HASH:
            create_backup(USERS_FILE)
            if os.path.exists(USERS_FILE):
                with open(USERS_FILE, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                with open(USERS_FILE, 'w', encoding='utf-8') as f:
                    for line in lines:
                        if line.strip() != str(user_id):
                            f.write(line)
            delete_user_from_excel(user_id, RESULTS_FILE)
            user_test_state.pop(user_id, None)
            user_mark[:] = [entry for entry in user_mark if entry['id'] != user_id]
            if user_id not in user_lang:
                user_lang[user_id] = 'ru'
            msg = {
                'tj': "✅ Тест аз нав оғоз шуд! Фармони /test ро истифода баред ва аз нав имтиҳон супоред.",
                'ru': "✅ Тест перезапущен! Используйте команду /test для повторной сдачи теста."
            }.get(user_lang[user_id], 'ru')
            sent_msg = await bot.send_message(user_id, msg)
            _track_for_deletion(user_id, sent_msg.message_id)
            if user_id not in all_message_ids:
                all_message_ids[user_id] = []
            all_message_ids[user_id].append(sent_msg.message_id)
        else:
            msg = {
                'tj': "❌ Парол нодуруст аст. Тест аз нав оғоз нашуд.",
                'ru': "❌ Неправильный пароль. Тест не был перезапущен."
            }.get(user_lang[user_id], 'ru')
            sent_msg = await bot.send_message(user_id, msg)
            _track_for_deletion(user_id, sent_msg.message_id)
            if user_id not in all_message_ids:
                all_message_ids[user_id] = []
            all_message_ids[user_id].append(sent_msg.message_id)
            try:
                await bot.delete_message(user_id, message.message_id)
            except Exception as e:
                logging.error(f"Error deleting password message {message.message_id} for user {user_id}: {e}")
            await send_pravial(message)
        password_waiting.pop(user_id, None)
        save_state()
    except Exception as e:
        logging.error(f"Error in check_password for user {user_id}: {e}")
        sent_msg = await bot.send_message(user_id, "Хато рух дод. Лутфан дубора кӯшиш кунед.")
        _track_for_deletion(user_id, sent_msg.message_id)
        if user_id not in all_message_ids:
            all_message_ids[user_id] = []
        all_message_ids[user_id].append(sent_msg.message_id)
# -----------------------------------------------------------------------------------



# ---------------------------------Интихоби забон-------------------------------------
async def language_selection(user_id: int):
    markup = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.add("Тоҷикӣ 🇹🇯", "Русский 🇷🇺")
    msg = await bot.send_message(
        user_id,
        "Пожалуйста, выберите ваш язык: 🇷🇺\nЛутфан забони худро интихоб намоед: 🇹🇯",
        reply_markup=markup
    )
    _track_for_deletion(user_id, msg.message_id)

@bot.message_handler(func=lambda message: message.text in ["Русский 🇷🇺", "Тоҷикӣ 🇹🇯"])
async def set_language(message):
    try:
        user_id = message.chat.id
        language_choice = message.text
        if language_choice == "Русский 🇷🇺":
            user_lang[user_id] = 'ru'
            sent = await bot.send_message(user_id, "Вы выбрали русский язык.", reply_markup=ReplyKeyboardRemove())
        elif language_choice == "Тоҷикӣ 🇹🇯":
            user_lang[user_id] = 'tj'
            sent = await bot.send_message(user_id, "Шумо забони тоҷикӣ - ро интихоб кардед.", reply_markup=ReplyKeyboardRemove())
        else:
            sent = await bot.send_message(
                user_id,
                "Инхел забон нест лутфан дубора кӯшиш кунед!\nТакого языка нет, пожалуйста, попробуйте еще раз!",
                reply_markup=ReplyKeyboardRemove()
            )
        _track_for_deletion(user_id, sent.message_id)
        await send_pravial(message)
        save_state()
    except Exception as e:
        logging.error(f"Error in set_language: {e}")
        await bot.send_message(user_id, "Хато рух дод. Лутфан дубора кӯшиш кунед.")

@bot.message_handler(commands=['language'])
async def lang(message):
    await language_selection(message.chat.id)
# -----------------------------------------------------------------------------------


# ---------------------------------Фармони /start------------------------------------
@bot.message_handler(commands=['start'])
async def start(message):
    try:
        user_id = message.chat.id
        first_name = message.from_user.first_name
        if user_id not in user_lang:
            await language_selection(user_id)
        else:
            msg = {
                'tj': (
                    f"Салом {first_name}, хуш омадед ба боти SoftClub test\n"
                    "ин бот барои муян кардани сатҳи дониши шумо мебошад! "
                    "Барои оғози тест фармони /test ро пахш кунед!"
                ),
                'ru': (
                    f"Здравствуйте {first_name}, добро пожаловать в бот SoftClub test!\n"
                    "Этот бот предназначен для определения вашего уровня знаний. "
                    "Для начала теста используйте команду /test!"
                )
            }.get(user_lang[user_id], 'ru')
            sent = await bot.send_message(user_id, msg)
            _track_for_deletion(user_id, sent.message_id)
        await _maybe_cancel_if_testing(user_id, reason="start_command")
        save_state()
    except Exception as e:
        logging.error(f"Error in start: {e}")
        await bot.send_message(message.chat.id, "Хато рух дод. Лутфан дубора кӯшиш кунед.")

async def send_pravial(message):
    try:
        user_id = message.chat.id
        if user_id not in user_lang:
            user_lang[user_id] = 'ru'
        msg = {
            'tj': 'Барои оғози кор бо бот дархости /start ро иҷро кунед.',
            'ru': 'Для начала работы с ботом выполните команду /start.'
        }.get(user_lang[user_id], 'ru')
        sent = await bot.send_message(user_id, msg)
        _track_for_deletion(user_id, sent.message_id)
        save_state()
    except Exception as e:
        logging.error(f"Error in send_pravial: {e}")
# -----------------------------------------------------------------------------------



# ---------------------------------Фармони /mark-------------------------------------
def get_user_score(user_id):
    try:
        for entry in user_mark:
            if entry['id'] == user_id:
                return entry['score']
        return None
    except Exception as e:
        logging.error(f"Error in get_user_score: {e}")
        return None


@bot.message_handler(commands=['mark'])
async def return_mark(message):
    try:
        user_id = message.chat.id
        first_name = message.from_user.first_name
        if user_id not in user_lang:
            user_lang[user_id] = 'ru'
        score = math.ceil(get_user_score(user_id)) if get_user_score(user_id) else None
        if score is not None:
            msg = {
                'tj': f'Баҳои ба даст овардаи шумо аз 100 бал --({score})-- бал мебошад!',
                'ru': f'Ваша оценка из 100 баллов: --({score})-- баллов!'
            }.get(user_lang[user_id], 'ru')
        else:
            msg = {
                'tj': f'{first_name}, шумо имтиҳон супоридагӣ нестед!',
                'ru': f'{first_name}, вы еще не сдавали тест!'
            }.get(user_lang[user_id], 'ru')
        sent = await bot.send_message(user_id, msg)
        _track_for_deletion(user_id, sent.message_id)
        save_state()
    except Exception as e:
        logging.error(f"Error in return_mark: {e}")
        await bot.send_message(user_id, "Хато рух дод. Лутфан дубора кӯшиш кунед.")
# -----------------------------------------------------------------------------------



# ---------------------------------Фармони /help-------------------------------------
@bot.message_handler(commands=['help'])
async def help_command(message):
    try:
        user_id = message.chat.id
        lang = user_lang.get(user_id, 'ru')
        help_text = {
            'tj': (
                "🤖 Ҳама фармонҳои бот SoftClub Академия:\n\n"
                "/start - Оғози кор бо бот\n"
                "/language - Тағйири забон\n"
                "/test - Оғози тест барои қабули донишҷӯён\n"
                "/mark - Дидани натиҷаи тест\n"
                "/restart_test - Такрори тест (барои иҷозатдодашудагон)\n"
                "/cancel_test - Бекор кардани тест (фаврӣ)\n"
                "/help - Ин паёми кӯмак\n\n"
                "Барои кӯмак ва маълумоти бештар бо маъмур тамос гиред."
            ),
            'ru': (
                "🤖 Все команды бота SoftClub Академия:\n\n"
                "/start - Начать работу с ботом\n"
                "/language - Сменить язык\n"
                "/test - Начать тест для приёма студентов\n"
                "/mark - Посмотреть результат теста\n"
                "/restart_test - Повтор теста (для авторизованных)\n"
                "/cancel_test - Немедленно отменить тест\n"
                "/help - Это сообщение помощи\n\n"
                "Для помощи и дополнительной информации свяжитесь с администратором."
            )
        }
        sent = await bot.send_message(user_id, help_text.get(lang, help_text['ru']))
        _track_for_deletion(user_id, sent.message_id)
        await _maybe_cancel_if_testing(user_id, reason="help_command")
        save_state()
    except Exception as e:
        logging.error(f"Error in help: {e}")
        await bot.send_message(user_id, "Хато рух дод. Лутфан дубора кӯшиш кунед.")
# -----------------------------------------------------------------------------------



# ---------------------------------Фармони /test-------------------------------------
def initialize_files():
    try:
        if not os.path.exists(USERS_FILE):
            with open(USERS_FILE, 'w', encoding='utf-8') as f:
                pass
        if not os.path.exists(RESULTS_FILE):
            wb = Workbook()
            ws = wb.active
            headers = ['ID', 'Имя', 'Фамилия', 'Username', 'Год рождения', 'Возраст', 'Баллы', 'Дата', 'Время']
            ws.append(headers)
            wb.save(RESULTS_FILE)
        logging.info("Files initialized successfully")
    except Exception as e:
        logging.error(f"Error initializing files: {e}")

initialize_files()
# -----------------------------------------------------------------------------------



# ---------------------------------Функсияҳо ----------------------------------------
async def delete_message_safe(chat_id, message_id):
    try:
        await bot.delete_message(chat_id, message_id)
    except Exception:
        pass

def _track_for_deletion(user_id: int, message_id: int):
    try:
        deletion_queue.append({
            'user_id': user_id,
            'message_id': message_id,
            'ts': datetime.now().isoformat()
        })
    except Exception as e:
        logging.error(f"_track_for_deletion error: {e}")

def _touch(user_id: int):
    try:
        last_activity[user_id] = datetime.now()
    except Exception as e:
        logging.error(f"_touch error: {e}")

async def cancel_test_and_cleanup(user_id: int, reason: str = "user_exit"):
    try:
        if user_id in user_test_state:
            # поккунии inline-тугмаҳо ва худи саволҳо
            for mid in user_test_state[user_id].get('message_ids', []):
                try:
                    await bot.edit_message_reply_markup(chat_id=user_id, message_id=mid, reply_markup=None)
                except Exception:
                    pass
                await delete_message_safe(user_id, mid)
            # дар хотима — сабт накардан ба Excel, холҳоро низ нест мекунем
            user_test_state.pop(user_id, None)
        # поккунии охирин паёмҳои ирсолшудаи ёрирасон (агар лозим донистед, монда мешавад)
        # ихтиёрӣ: ба корбар паёми огоҳӣ
        lang = user_lang.get(user_id, 'ru')
        msg = {
            'tj': "❌ Тест бекор карда шуд. Ҳама саволҳо ва натиҷаҳо пок гардиданд. Барои оғози нав /test.",
            'ru': "❌ Тест отменён. Все вопросы и результаты удалены. Для нового старта — /test."
        }.get(lang, 'ru')
        notice = await bot.send_message(user_id, msg)
        _track_for_deletion(user_id, notice.message_id)
        # ҳамзамон корбарро аз USERS_FILE бардорем, то битавонад аз нав супорад
        if os.path.exists(USERS_FILE):
            with open(USERS_FILE, 'r', encoding='utf-8') as f:
                lines = [ln for ln in f if ln.strip() != str(user_id)]
            with open(USERS_FILE, 'w', encoding='utf-8') as f:
                f.writelines(lines)
        # аз рӯйхати баҳогирӣ низ пок кунем
        global user_mark
        user_mark = [m for m in user_mark if m.get('id') != user_id]
        save_state()
        logging.info(f"Test cancelled for {user_id}. Reason={reason}")
    except Exception as e:
        logging.error(f"cancel_test_and_cleanup error: {e}")

async def _maybe_cancel_if_testing(user_id: int, reason: str):
    try:
        st = user_test_state.get(user_id)
        if st and st.get('step') == 'testing':
            await cancel_test_and_cleanup(user_id, reason=reason)
    except Exception as e:
        logging.error(f"_maybe_cancel_if_testing error: {e}")

async def cleanup_old_messages_task():
    while True:
        try:
            now = datetime.now()
            keep = []
            for item in deletion_queue:
                try:
                    ts = datetime.fromisoformat(item['ts'])
                    if (now - ts).total_seconds() >= QUESTION_TTL_SECONDS:
                        await delete_message_safe(item['user_id'], item['message_id'])
                    else:
                        keep.append(item)
                except Exception:
                    keep.append(item)
            if len(keep) != len(deletion_queue):
                deletion_queue[:] = keep
                save_state()
        except Exception as e:
            logging.error(f"cleanup_old_messages_task error: {e}")
        await asyncio.sleep(600)  # ҳар 10 дақиқа
# ========================================================


async def send_question(user_id, question_index):
    try:
        if user_id not in user_test_state:
            logging.error(f"User {user_id} not in user_test_state during send_question")
            return
        lang = user_lang.get(user_id, 'ru')
        test_set = test_tjk if lang == 'tj' else test_rus
        if question_index >= len(test_set):
            await end_test(user_id)
            return
        # лимити умумӣ
        if (datetime.now() - user_test_state[user_id]['start_time']).total_seconds() > TEST_TOTAL_LIMIT_SECONDS:
            await end_test(user_id)
            return
        question = test_set[question_index]
        markup = InlineKeyboardMarkup()
        markup.row(
            InlineKeyboardButton(question['options'][0], callback_data=f"ans_{user_id}_{question_index}_A"),
            InlineKeyboardButton(question['options'][1], callback_data=f"ans_{user_id}_{question_index}_B")
        )
        markup.row(
            InlineKeyboardButton(question['options'][2], callback_data=f"ans_{user_id}_{question_index}_C"),
            InlineKeyboardButton(question['options'][3], callback_data=f"ans_{user_id}_{question_index}_D")
        )
        # Нест накардани саволҳои қаблӣ, то саволҳо ва тугмаҳо боқӣ монанд
        if os.path.exists(question['image']):
            with open(question['image'], 'rb') as photo:
                msg = await bot.send_photo(
                    user_id,
                    photo,
                    caption=question['question'],
                    reply_markup=markup
                )
        else:
            msg = await bot.send_message(
                user_id,
                question['question'],
                reply_markup=markup
            )
        # === ADDED: пайгирӣ ва last_q_msg_id ===
        if 'message_ids' not in user_test_state[user_id]:
            user_test_state[user_id]['message_ids'] = []
        user_test_state[user_id]['message_ids'].append(msg.message_id)
        user_test_state[user_id]['current_question'] = question_index
        user_test_state[user_id]['last_q_msg_id'] = msg.message_id  # барои бастани тугмаҳо баъд аз ҷавоб
        _track_for_deletion(user_id, msg.message_id)
        _touch(user_id)
        # =========================================
        logging.info(f"Sent question {question_index} to user {user_id}, message_id: {msg.message_id}")
        save_state()
    except Exception as e:
        logging.error(f"Error in send_question for user {user_id}: {e}")
        error_msg = {
            'tj': f"Хато: {str(e)}. Лутфан дубора кӯшиш кунед!",
            'ru': f"Ошибка: {str(e)}. Попробуйте снова!"
        }.get(lang, 'ru')
        msg = await bot.send_message(user_id, error_msg)
        _track_for_deletion(user_id, msg.message_id)
        await end_test(user_id)

async def delete_previous_question(user_id):
    try:
        pass
    except Exception as e:
        logging.error(f"Error in delete_previous_question: {e}")

async def save_results(user_id):
    try:
        if user_id not in user_test_state:
            logging.error(f"User {user_id} not in user_test_state during save_results")
            return
        user_data = user_test_state[user_id]
        current_year = datetime.now().year
        age = current_year - user_data['birth_year']
        create_backup(RESULTS_FILE)
        wb = load_workbook(RESULTS_FILE)
        ws = wb.active
        ws.append([
            user_id,
            user_data.get('first_name', ''),
            user_data.get('last_name', ''),
            user_data.get('username', ''),
            user_data.get('birth_year', ''),
            age,
            math.ceil(user_data.get('score', 0)),
            datetime.now().strftime("%Y-%m-%d"),
            datetime.now().strftime("%H:%M:%S")
        ])
        wb.save(RESULTS_FILE)
        logging.info(f"Results saved for user {user_id}")
    except Exception as e:
        logging.error(f"Error saving results for {user_id}: {e}")

async def end_test(user_id):
    try:
        if user_id not in user_test_state:
            logging.error(f"User {user_id} not in user_test_state during end_test")
            return
        await save_results(user_id)
        lang = user_lang.get(user_id, 'ru')
        test_set = test_tjk if lang == 'tj' else test_rus
        raw_score = user_test_state[user_id].get('score', 0)
        max_score = 100
        normalized_score = math.ceil(raw_score)
        result_msg = {
            'tj': f"📊 Тест ба охир расид!\nБаҳои шумо: {normalized_score} аз {max_score} бал.",
            'ru': f"📊 Тест завершен!\nВаш результат: {normalized_score} из {max_score} баллов."
        }.get(lang, 'ru')
        if 'message_ids' in user_test_state[user_id]:
            for msg_id in user_test_state[user_id]['message_ids']:
                try:
                    await bot.edit_message_reply_markup(
                        chat_id=user_id,
                        message_id=msg_id,
                        reply_markup=None
                    )
                    logging.info(f"Removed inline keyboard for message {msg_id} of user {user_id}")
                except Exception as e:
                    logging.error(f"Error removing inline keyboard for message {msg_id} of user {user_id}: {e}")
        sent = await bot.send_message(user_id, result_msg)
        _track_for_deletion(user_id, sent.message_id)
        user_mark.append({'id': user_id, 'score': normalized_score})
        del user_test_state[user_id]
        logging.info(f'Test ended for user {user_id}, score: {normalized_score}')
        save_state()
    except Exception as e:
        logging.error(f"Error in end_test for user {user_id}: {e}")
        await bot.send_message(user_id, "Хато рух дод. Лутфан дубора кӯшиш кунед.")

async def send_monthly_report():
    while True:
        try:
            today = datetime.now()
            if today.day == 5 and os.path.exists(RESULTS_FILE):
                for _ in range(3):  
                    try:
                        wb = load_workbook(RESULTS_FILE)
                        ws = wb.active
                        rows = list(ws.iter_rows(min_row=2, values_only=True))
                        scores = [r[6] for r in rows if r and isinstance(r[6], (int, float))]
                        total = len(scores)
                        avg = round(sum(scores)/total, 2) if total else 0
                        best = max(scores) if scores else 0
                        month_name = today.strftime("%B")
                        year = today.year
                        caption = f"🗓️ Отчет за {month_name} {year}\n👥 Кол-во: {total}\n📈 Средний балл: {avg}\n🏆 Максимум: {best}"
                        with open(RESULTS_FILE, 'rb') as report:
                            for admin_id in [ADMIN_ID, ADMIN3_ID, ADMIN4_ID]:
                                await bot.send_document(admin_id, report, caption=caption)
                        wb_new = Workbook()
                        ws_new = wb_new.active
                        ws_new.append(['ID', 'Имя', 'Фамилия', 'Username', 'Год рождения', 'Возраст', 'Баллы', 'Дата', 'Время'])
                        wb_new.save(RESULTS_FILE)
                        open(USERS_FILE, 'w', encoding='utf-8').close()
                        logging.info("Monthly report sent successfully")
                        break
                    except Exception as e:
                        logging.error(f"Error sending report: {e}")
                        await asyncio.sleep(60)
            await asyncio.sleep(43200) 
        except Exception as e:
            logging.error(f"Error in send_monthly_report: {e}")
            await asyncio.sleep(60)
# -----------------------------------------------------------------------------------



# ---------------------------------Командаи тест-------------------------------------
@bot.message_handler(commands=['test'])
async def start_test(message):
    try:
        user_id = message.chat.id
        username = message.from_user.username or "N/A"
        _touch(user_id)
        if user_id not in user_lang:
            await language_selection(user_id)
            return
        if user_id in user_test_state and user_test_state[user_id].get('step') == 'testing':
            lang = user_lang.get(user_id, 'ru')
            msg = {
                'tj': "Шумо як тести ниматамом доред. Оё мехоҳед аз саволи охирин идома диҳед?",
                'ru': "У вас есть незавершенный тест. Хотите продолжить с последнего вопроса?"
            }.get(lang, 'ru')
            markup = InlineKeyboardMarkup()
            markup.add(
                InlineKeyboardButton("Бале/Да", callback_data=f"resume_test_{user_id}"),
                InlineKeyboardButton("Не/Нет", callback_data=f"restart_test_{user_id}")
            )
            sent = await bot.send_message(user_id, msg, reply_markup=markup)
            _track_for_deletion(user_id, sent.message_id)
            return
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            if str(user_id) in f.read():
                msg = {
                    'tj': "Шумо аллакай тестроҳоро анҷом додаед!",
                    'ru': "Вы уже прошли тест!"
                }.get(user_lang[user_id], 'ru')
                sent = await bot.send_message(user_id, msg)
                _track_for_deletion(user_id, sent.message_id)
                return
        user_test_state[user_id] = {
            'step': 'ask_full_name',
            'username': username,
            'first_name': message.from_user.first_name
        }
        current_year = datetime.now().year
        msg = {
            'tj': "Лутфан ном ва насаби худро фиристед (масалан: Али Алиев):",
            'ru': "Пожалуйста, отправьте ваше имя и фамилию (например: Иван Иванов):"
        }.get(user_lang[user_id], 'ru')
        sent = await bot.send_message(user_id, msg)
        _track_for_deletion(user_id, sent.message_id)
        save_state()
    except Exception as e:
        logging.error(f"Error in start_test: {e}")
        await bot.send_message(user_id, "Хато рух дод. Лутфан дубора кӯшиш кунед.")

@bot.message_handler(func=lambda m: user_test_state.get(m.chat.id, {}).get('step') == 'ask_full_name')
async def process_name(message):
    try:
        user_id = message.chat.id
        _touch(user_id)
        parts = message.text.strip().split()
        if len(parts) < 2:
            msg = {
                'tj': "Лутфан ном ва насаби худро дуруст дохил кунед!",
                'ru': "Пожалуйста, введите имя и фамилию правильно!"
            }.get(user_lang[user_id], 'ru')
            sent = await bot.send_message(user_id, msg)
            _track_for_deletion(user_id, sent.message_id)
            return
        user_test_state[user_id].update({
            'first_name': parts[0],
            'last_name': ' '.join(parts[1:]),
            'step': 'ask_birth_year'
        })
        current_year = datetime.now().year
        msg = {
            'tj': f"Лутфан соли таваллуди худро фиристед (масалан: {current_year-20}):",
            'ru': f"Пожалуйста, отправьте ваш год рождения (например: {current_year-20}):"
        }.get(user_lang[user_id], 'ru')
        sent = await bot.send_message(user_id, msg)
        _track_for_deletion(user_id, sent.message_id)
        save_state()
    except Exception as e:
        logging.error(f"Error in process_name: {e}")
        await bot.send_message(user_id, "Хато рух дод. Лутфан дубора кӯшиш кунед.")

@bot.message_handler(func=lambda m: user_test_state.get(m.chat.id, {}).get('step') == 'ask_birth_year')
async def process_birth_year(message):
    try:
        user_id = message.chat.id
        _touch(user_id)
        birth_year = int(message.text.strip())
        current_year = datetime.now().year
        if not 1900 <= birth_year <= current_year:
            raise ValueError
        with open(USERS_FILE, 'a', encoding='utf-8') as f:
            f.write(f"{user_id}\n")
        user_test_state[user_id].update({
            'birth_year': birth_year,
            'step': 'testing',
            'current_question': 0,
            'score': 0,
            'start_time': datetime.now(),
            'message_ids': []
        })
        await send_question(user_id, 0)
        save_state()
    except ValueError:
        msg = {
            'tj': "Лутфан соли таваллудатонро дуруст ворид кунед!",
            'ru': "Пожалуйста, введите корректный год рождения!"
        }.get(user_lang[user_id], 'ru')
        sent = await bot.send_message(user_id, msg)
        _track_for_deletion(user_id, sent.message_id)
    except Exception as e:
        logging.error(f"Error in process_birth_year: {e}")
        await bot.send_message(user_id, "Хато рух дод. Лутфан дубора кӯшиш кунед.")

@bot.callback_query_handler(func=lambda call: call.data.startswith('ans_'))
async def handle_answer(call):
    try:
        logging.info(f"Received callback: {call.data}")
        _, user_id, question_index, answer = call.data.split('_')
        user_id = int(user_id)
        question_index = int(question_index)
        _touch(user_id)
        if user_id not in user_test_state:
            logging.error(f"User {user_id} not in user_test_state")
            await bot.answer_callback_query(call.id, "Тест завершен!")
            return
        if user_test_state[user_id]['current_question'] != question_index:
            logging.warning(f"Question mismatch for user {user_id}: expected {user_test_state[user_id]['current_question']}, got {question_index}")
            msg = {
                'tj': "Ин савол аллакай коркарда шудааст!",
                'ru': "Этот вопрос уже обработан!"
            }.get(user_lang.get(user_id, 'ru'), 'ru')
            await bot.answer_callback_query(call.id, msg)
            return
        if (datetime.now() - user_test_state[user_id]['start_time']).total_seconds() > TEST_TOTAL_LIMIT_SECONDS:
            logging.info(f"Test timed out for user {user_id}")
            await end_test(user_id)
            await bot.answer_callback_query(call.id)
            return
        lang = user_lang.get(user_id, 'ru')
        test_set = test_tjk if lang == 'tj' else test_rus

        last_q = user_test_state[user_id].get('last_q_msg_id')
        if last_q:
            try:
                await bot.edit_message_reply_markup(chat_id=user_id, message_id=last_q, reply_markup=None)
            except Exception as e:
                logging.error(f"edit_message_reply_markup fail q={last_q}: {e}")

        is_correct = answer == test_set[question_index]['correct']
        if is_correct:
            user_test_state[user_id]['score'] += (100 / len(test_set))
            result_msg = {
                'tj': "✅ Ҷавоби дуруст!",
                'ru': "✅ Правильный ответ!"
            }.get(lang, 'ru')
        else:
            result_msg = {
                'tj': f"❌ Ҷавоби нодуруст! Ҷавоби дуруст: {test_set[question_index]['options'][ord(test_set[question_index]['correct']) - ord('A')]}",
                'ru': f"❌ Неправильный ответ! Правильный ответ: {test_set[question_index]['options'][ord(test_set[question_index]['correct']) - ord('A')]}"
            }.get(lang, 'ru')
        sent = await bot.send_message(user_id, result_msg)
        _track_for_deletion(user_id, sent.message_id)
        # Гузариш ба саволи навбатӣ
        await send_question(user_id, question_index + 1)
        await bot.answer_callback_query(call.id, "Ҷавоб қабул шуд!")
        save_state()
    except Exception as e:
        logging.error(f"Error in handle_answer for user {user_id}: {e}")
        error_msg = {
            'tj': f"Хато дар коркарди ҷавоб: {str(e)}!",
            'ru': f"Ошибка в обработке ответа: {str(e)}!"
        }.get(user_lang.get(user_id, 'ru'), 'ru')
        try:
            await bot.answer_callback_query(call.id, error_msg)
        except Exception:
            pass

@bot.callback_query_handler(func=lambda call: call.data.startswith('resume_test_'))
async def resume_test(call):
    try:
        user_id = int(call.data.split('_')[-1])
        _touch(user_id)
        if user_id in user_test_state and user_test_state[user_id].get('step') == 'testing':
            await send_question(user_id, user_test_state[user_id]['current_question'])
            await bot.answer_callback_query(call.id, "Тест идома дода шуд!")
        save_state()
    except Exception as e:
        logging.error(f"Error in resume_test: {e}")
        await bot.answer_callback_query(call.id, "Хато рух дод.")

@bot.callback_query_handler(func=lambda call: call.data.startswith('restart_test_'))
async def restart_test(call):
    try:
        user_id = int(call.data.split('_')[-1])
        _touch(user_id)
        if user_id in user_test_state:
            del user_test_state[user_id]
        await start_test(call.message)
        await bot.answer_callback_query(call.id, "Тест аз нав оғоз шуд!")
        save_state()
    except Exception as e:
        logging.error(f"Error in restart_test: {e}")
        await bot.answer_callback_query(call.id, "Хато рух дод.")
# -----------------------------------------------------------------------------------


@bot.message_handler(commands=['cancel_test'])
async def user_cancel_test(message):
    user_id = message.chat.id
    await cancel_test_and_cleanup(user_id, reason="user_cancel_command")


@bot.message_handler(func=lambda m: user_test_state.get(m.chat.id, {}).get('step') == 'testing' and m.text and m.text.startswith('/'))
async def cancel_on_any_command_in_testing(message):
    user_id = message.chat.id
    await cancel_test_and_cleanup(user_id, reason="other_command_during_testing")


async def watchdog_task():
    while True:
        try:
            now = datetime.now()
            for uid, st in list(user_test_state.items()):
                if st.get('step') == 'testing':
                    la = last_activity.get(uid, st.get('start_time', now))
                    if (now - la).total_seconds() > WATCHDOG_TIMEOUT_SECONDS:
                        await cancel_test_and_cleanup(uid, reason="inactivity_watchdog")
            await asyncio.sleep(60)
        except Exception as e:
            logging.error(f"watchdog_task error: {e}")
            await asyncio.sleep(60)
# ========================================================


# ---------------------------------Оғози бот-------------------------------------
async def start_bot():
    max_retries = 5
    retry_count = 0
    while retry_count < max_retries:
        try:
            await bot.infinity_polling()
        except Exception as e:
            retry_count += 1
            logging.error(f"[ERROR] Reconnecting in 5s... {e}")
            await asyncio.sleep(5)
        if retry_count == max_retries:
            logging.critical("Max retries reached. Stopping bot.")
            break

async def main():
    try:
        await set_commands()
        asyncio.create_task(save_state_periodically())
        asyncio.create_task(send_monthly_report())
        asyncio.create_task(cleanup_old_messages_task())
        asyncio.create_task(watchdog_task())
        await start_bot()
    except Exception as e:
        logging.error(f"Main loop error: {e}")

if __name__ == '__main__':
    asyncio.run(main())
